from odoo import api, fields, models, _
import base64
import io
import logging
from datetime import datetime

from odoo.tools.translate import _, pycompat
from odoo.exceptions import UserError
from xlrd import open_workbook
from openpyxl import Workbook
import openpyxl

from openpyxl import load_workbook

_logger = logging.getLogger(__name__)


class MeterReadingImport(models.TransientModel):
    _name = 'meter.reading.import'
    _description = 'Import Meter Reading'

    data_file = fields.Binary('Meter Reading File', required=True,
                              help='Select your Meter Reading csv file here - make sure it is the latest one !.')
    input_layout = fields.Selection([('fm', "FM Audit"), ("man", "Manual"),("avg","Average"),("his","History")], "Choose the File Format for this input",
                                    default='fm')
    def update_readings(self, black, colour):
        print(black, colour)
        return

    def import_readings(self):
        # line_obj =  self.env['sale.subscription.line']
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M")
        serial_no = ''
        lines = []

        error_obj = self.env['meter.reading.error']
        error_obj.search([]).unlink()  # Delete All prevoius messages
        line_obj = self.env['sale.subscription.line']
        csv_data = base64.b64decode(self.data_file)
        csv_input_file = pycompat.csv_reader(io.BytesIO(csv_data), quotechar='"', delimiter=',')
        err = 'Starting Import  ' + dt_string
        error_obj.create({'name': err})
        lines = []

        for row in csv_input_file:
            serial_no = row[0].strip('"')
            colour = row[1]
            black = row[2]
            line_ids = line_obj.search([('x_serial_number_id', '=', serial_no)])
            if line_ids:
                for line in line_ids:
                    if black.isdigit():
                        if line.name == 'Black copies' and int(black) > 0:
                            print('found Black copies serialnumber ', serial_no, line.name,
                                  line.analytic_account_id.name)
                            line.write({'x_copies_last': black})
                    if colour.isdigit():
                        if line.name == 'Colour copies' and int(colour) > 0:
                            print('found Colour copies serialnumber ', serial_no, line.name,
                                  line.analytic_account_id.name)
                            line.write({'x_copies_last': colour})
                    if self.input_layout == 'fm':
                        line.write({'x_reading_type_last': 'FM Audit'})
                    if self.input_layout == 'man':
                        line.write({'x_reading_type_last': 'Manual'})
                    if self.input_layout == 'avg':
                        line.write({'x_reading_type_last': 'Average'})

        print('Finished')
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M")
        message = dt_string + ' Import Finished - last record was ' + serial_no
        error_obj.create({'name': message})
        return {
            'view_type': 'form',
            'view_mode': 'tree',
            'res_model': 'meter.reading.error',
            'target': 'current',
            'res_id': 121,
            'type': 'ir.actions.act_window'
        }

    def read_xls_import(self):
        now = datetime.now()
        dt_string = now.strftime("%d/%m/%Y %H:%M")
        error_obj = self.env['meter.reading.error']
        error_obj.search([]).unlink()  # Delete All prevoius messages
        line_obj = self.env['sale.subscription.line']
        serial_no = ''
        line_num = 0
        message = ''
        contact_no = ''

        logging.warning("Data file is %s", self.data_file)

        # Read xls file
        wb = openpyxl.load_workbook(

            filename=io.BytesIO(base64.b64decode(self.data_file)), read_only=True
        )
        ws = wb.active

        # If the reading type is history we dont want to import readings into subscription lines
        # We only want to populate the history table as it has no bearing on current or future readings


        if self.input_layout == 'his1':
            logging.warning("Doing History meter reading import")
            for row in ws.iter_rows(min_row=2, max_row=None, min_col=None,

                                    max_col=None, values_only=True):
                if row[0] == '':
                    continue
                contract_id = ''
                serial_no_id = ''
                product_id = ''
                machine_id = ''
                copies_previous = ''
                copies_last = ''
                no_of_copies = ''
                reading_type = 'Average'

                # Lookup the contract and find the serial number associated with the contract
                logging.warning("Contract and serial number is %s %s %s %s", row[0], row[1],row[7], row[2])

                contract = self.env['sale.subscription'].search([('name','=',row[7])])
                if contract:
                    for machine in contract.x_machine_ids:

                        if machine.name == row[2] and machine.x_main_product:
                            logging.warning("Found the contract and Machine %s %s %s", contract.name, machine.name, row[2])
                            logging.warning("Main Machine and ID is %s %s", machine.name, machine.id)
                            contract_id = contract.id
                            product_id = self.env['product.product'].search([('name','=', row[3])]).id
                            serial_no_id = machine.id
                            machine_id = machine.product_id.id
                            copies_previous = row[4]
                            copies_last = row[5]
                            no_of_copies = row[6]
                            reading_type = reading_type

                            # First check if a hostory record of the same values exist and either delete
                            # or raise an error condition
                            logging.warning("Product id is %s", product_id)
                            duplicate_rec = self.env['meter.reading.history'].search([('contract_id','=',contract_id),
                                                                                      ('serial_no_id','=',serial_no_id),
                                                                                      ('product_id','=', product_id),
                                                                                      ('machine_id','=',machine_id),
                                                                                      ('copies_previous','=',copies_previous),
                                                                                      ('copies_last','=',copies_last),
                                                                                      ('reading_type','=',reading_type)])
                            if duplicate_rec:
                                raise UserError(_("METER READING ALREADY EXIST\n\n"
                                                  "Contract: %s \n"
                                                  "Machine: %s \n"
                                                  "Serial Number: %s \n"
                                                  "Product: %s \n"
                                                  "Date Imported: %s \n"
                                                  "Exception: 1. Check the Meter Reading History report given the above data .\n") %
                                                (contract_id, duplicate_rec.machine_id.name,
                                                 duplicate_rec.serial_no_id.name,duplicate_rec.product_id.name,
                                                 duplicate_rec.create_date))


                            # Create a history record
                            history_rec = self.env['meter.reading.history'].create(
                                {
                                    'name': contract.partner_id.name,
                                    'create_date': row[0],
                                    'contract_id': contract_id,
                                    'product_id': product_id,
                                    'serial_no_id': serial_no_id,
                                    'machine_id': machine_id,
                                    'copies_previous': copies_previous,
                                    'copies_last': copies_last,
                                    'no_of_copies':no_of_copies,
                                    'reading_type':reading_type,
                                })
                            if history_rec:
                                logging.warning("History record created %s", history_rec)
                            else:
                                logging.warning("No history record created")

                    # Let's do the main machines first
                    #serial_no = self.env['stock.production.lot'].search()
                    # raise UserError(
                    #     _('Contract found'))
                else:
                    logging.warning("No Contract found %s", row[11])
                    raise UserError(
                        _('No Contract found'))

        if self.input_layout == 'his':
            logging.warning("Doing History meter reading import")
            for row in ws.iter_rows(min_row=2, max_row=None, min_col=None,

                                    max_col=None, values_only=True):
                if row[0] == '':
                    continue
                contract_id = ''
                serial_no_id = ''
                product_id = ''
                machine_id = ''
                copies_previous = ''
                copies_last = ''
                no_of_copies = ''
                copies_black = ''
                copies_colour = ''
                reading_type = 'History'

                # Lookup the contract and find the serial number associated with the contract
                #logging.warning("Contract and serial number is %s %s %s %s", row[0], row[1],row[7], row[2])
                logging.warning("Contract is %s", row[0])
                contract = self.env['sale.subscription'].search([('name','=',row[0])])
                if contract:
                    for machine in contract.x_machine_ids:
                        if machine.x_main_product:
                            logging.warning("Found the contract and Machine %s %s", contract.name, machine.name)
                            logging.warning("Main Machine and ID is product code is %s %s %s", machine.name, machine.id, row[6])
                            contract_id = contract.id
                            #product_id = self.env['product.product'].search([('name','=', row[3])]).id
                            serial_no_id = machine.id
                            machine_id = machine.product_id.id

                            # Check if there is any copies on the contract before writing historical data
                            if row[6].strip() == 'COL':
                                logging.warning("Checking colour copies")
                                product_id = self.env['product.template'].search([('name','=', 'Colour copies')]).id
                                if product_id:
                                    logging.warning("===The product is %s", product_id)
                                else:
                                    logging.warning("===Could not find colour copies")
                            if row[6].strip() == 'BLK':
                                logging.warning("Checking black copies")
                                product_id = self.env['product.template'].search([('name','=', 'Black copies')]).id
                                if product_id:
                                    logging.warning("=The product is %s", product_id)
                                else:
                                    logging.warning("=Could not find black copies")

                            copies_previous = row[11]
                            copies_last = row[9]
                            no_of_copies = row[12]

                            reading_type = reading_type

                            # First check if a hostory record of the same values exist and either delete
                            # or raise an error condition
                            logging.warning("Product id is %s", product_id)
                            duplicate_rec = self.env['meter.reading.history'].search([('contract_id','=',contract_id),
                                                                                      ('serial_no_id','=',serial_no_id),
                                                                                      ('product_id','=', product_id),
                                                                                      ('machine_id','=',machine_id),
                                                                                      ('copies_previous','=',copies_previous),
                                                                                      ('copies_last','=',copies_last),
                                                                                      ('reading_type','=',reading_type)])
                            if duplicate_rec:
                                raise UserError(_("METER READING ALREADY EXIST\n\n"
                                                  "Contract: %s \n"
                                                  "Machine: %s \n"
                                                  "Serial Number: %s \n"
                                                  "Product: %s \n"
                                                  "Date Imported: %s \n"
                                                  "Exception: 1. Check the Meter Reading History report given the above data .\n") %
                                                (contract_id, duplicate_rec.machine_id.name,
                                                 duplicate_rec.serial_no_id.name,duplicate_rec.product_id.name,
                                                 duplicate_rec.create_date))


                            # Create a history record
                            history_rec = self.env['meter.reading.history'].create(
                                {
                                    'name': contract.partner_id.name,
                                    'create_date': row[0],
                                    'contract_id': contract_id,
                                    'product_id': product_id,
                                    'serial_no_id': serial_no_id,
                                    'machine_id': machine_id,
                                    'copies_previous': copies_previous,
                                    'copies_last': copies_last,
                                    'no_of_copies':no_of_copies,
                                    'reading_type':reading_type,
                                })
                            if history_rec:
                                logging.warning("History record created %s", history_rec)
                            else:
                                logging.warning("No history record created")

                    # Let's do the main machines first
                    #serial_no = self.env['stock.production.lot'].search()
                    # raise UserError(
                    #     _('Contract found'))
                else:
                    logging.warning("No Contract found %s", row[0])
                    message = dt_string + ' Error: ' + ' Contract: ' + str(row[0]) + ' Not found in Odoo'
                    error_obj.create({'name': message})

        if self.input_layout == 'fm':
            logging.warning("FM Audit import")
            logging.warning("FM AUDIT IMPORT %s", ws)
            for row in ws.iter_rows(min_row=2, max_row=None, min_col=None,

                                    max_col=None, values_only=True):
                logging.warning("Record %s %s %s", type(row[0]),type(row[1]),type(row[2]))

                if row[0] == '':
                    continue

                serial_no = row[0]

                colour = row[1]
                black = row[2]
                # if isinstance(serial_no, int):
                #     serial_no = str(serial_no)
                #     logging.warning("Serial number is %s %s %s", serial_no, type(serial_no), black)
                # logging.warning("Reading line %s %s %s %s", serial_no, black, colour, str(line_num))

                line_ids = line_obj.search([('x_serial_number_id', '=', serial_no)])
                if line_ids:
                    for line in line_ids:
                        # logging.warning("line name is %s %s %s %s", serial_no, line.name, black, colour)
                        if line.name == 'Black copies' and int(black) > 0:
                            # logging.warning('found Black copies serial number %s %s %s', serial_no, line.name,
                            #       line.analytic_account_id.name)
                            line.write({'x_copies_last': black})
                            message = dt_string + ' Imported ' + 'Serial No: ' + serial_no + ' Line No' + str(
                                line_num) + ' Black =  ' + str(black) + ' x_copies_last = ' + str(line.x_copies_last)
                            error_obj.create({'name': message})
                        if line.name == 'Colour copies' and int(colour) > 0:
                            # logging.warning('found Colour copies serial number %s %s %s', serial_no, line.name,
                            #       line.analytic_account_id.name)
                            line.write({'x_copies_last': colour})
                            message = dt_string + ' Imported ' + 'Serial No: ' + serial_no + ' Line No=' + str(
                                line_num) + ' Colour =  ' + str(colour) + ' x_copies_last = ' + str(line.x_copies_last)
                            error_obj.create({'name': message})

                        if self.input_layout == 'fm':
                            line.write({'x_reading_type_last': 'FM Audit'})
                        if self.input_layout == 'man':
                            line.write({'x_reading_type_last': 'Manual'})
                        if self.input_layout == 'avg':
                            line.write({'x_reading_type_last': 'Average'})
                        line.analytic_account_id.sudo().message_post(body=message)

                line_num += 1

            # Run an audit on each line that was processed to ascertain that imported readings match the contents of the file
            for row in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                    max_col=None, values_only=True):
                if row[0] == '':
                    continue
                line_num += 1
                serial_no = row[0]
                colour = row[1]
                black = row[2]
                line_ids = line_obj.search([('x_serial_number_id', '=', serial_no)])
                if line_ids:
                    for line in line_ids:
                        logging.warning("Working with serial number - %s %s %s", serial_no,colour, black )
                        if line.name == 'Black copies' and line.x_copies_last != int(black):
                            message = dt_string + 'Error' + ' Serial No: ' + serial_no + 'Line No' + str(
                                line_num) + ' Black =  ' + str(black) + ' x_copies_last = ' + str(line.x_copies_last)
                            error_obj.create({'name': message})

                        if line.name == 'Colour copies' and line.x_copies_last != int(colour):
                            message = dt_string + 'Error' + ' Serial No: ' + serial_no + 'Line No' + str(
                                line_num) + ' Colour =  ' + str(colour) + ' x_copies_last = ' + str(line.x_copies_last)
                            error_obj.create({'name': message})

        print('Finished')
        message = dt_string + ' Import Finished - last record was ' + serial_no + 'Line No' + str(line_num)
        error_obj.create({'name': message})
        return {
            'view_type': 'form',
            'view_mode': 'tree',
            'res_model': 'meter.reading.error',
            'target': 'current',
            'res_id': 121,
            'type': 'ir.actions.act_window'
        }


class MeterReadingHistory(models.Model):
    _name = 'meter.reading.history'
    _description = 'Meter Reading History'
    _order = 'create_date desc, name, contract_id'

    name = fields.Char(string='Customer')
    contract_id = fields.Many2one('sale.subscription', string='Contract')
    product_id = fields.Many2one('product.product', string='Product')
    serial_no_id = fields.Many2one('stock.production.lot', 'Serial Number')
    machine_id = fields.Many2one('product.product', related='serial_no_id.product_id', string='Machine')
    copies_previous = fields.Integer('Previous Reading')
    copies_last = fields.Integer('Last Reading')
    no_of_copies = fields.Integer('Qty')
    reading_type = fields.Char('Reading Type')

    def _close_window(self):
        return {'type': 'ir.actions.act_window_close'}
